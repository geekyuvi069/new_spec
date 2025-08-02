import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TraceabilityMatrix:
    def __init__(self):
        self.matrix_data = None
    
    def generate_matrix(self, requirements, test_cases):
        """
        Generate traceability matrix mapping requirements to test cases.
        """
        # Create matrix structure
        matrix = {
            'requirements': [],
            'test_cases': [],
            'mappings': {},
            'metadata': {
                'generation_date': datetime.now().isoformat(),
                'total_requirements': len(requirements),
                'total_test_cases': len(test_cases)
            }
        }
        
        # Process requirements
        for req in requirements:
            req_entry = {
                'id': req['id'],
                'type': req.get('type', 'general'),
                'priority': req.get('priority', 'medium'),
                'category': req.get('category', 'general'),
                'content': req['content'][:200] + "..." if len(req['content']) > 200 else req['content'],
                'covered': False,
                'test_case_count': 0
            }
            matrix['requirements'].append(req_entry)
        
        # Process test cases and create mappings
        for tc in test_cases:
            tc_entry = {
                'id': tc.get('id', 'Unknown'),
                'title': tc.get('title', 'Untitled'),
                'type': tc.get('type', 'Functional'),
                'priority': tc.get('priority', 'Medium'),
                'status': tc.get('status', 'Generated'),
                'requirement_mappings': []
            }
            
            # Find mapped requirements for this test case
            mapped_req_ids = self._find_mapped_requirements(tc, requirements)
            tc_entry['requirement_mappings'] = mapped_req_ids
            
            # Update requirement coverage
            for req_id in mapped_req_ids:
                if req_id not in matrix['mappings']:
                    matrix['mappings'][req_id] = []
                matrix['mappings'][req_id].append(tc.get('id'))
                
                # Update requirement coverage status
                for req in matrix['requirements']:
                    if req['id'] == req_id:
                        req['covered'] = True
                        req['test_case_count'] += 1
                        break
            
            matrix['test_cases'].append(tc_entry)
        
        self.matrix_data = matrix
        return matrix
    
    def _find_mapped_requirements(self, test_case, requirements):
        """
        Find requirements mapped to a test case.
        """
        mapped_reqs = []
        
        # Check explicit requirement mapping
        explicit_req = test_case.get('requirement_id')
        if explicit_req:
            mapped_reqs.append(explicit_req)
        
        # Use semantic/keyword matching for additional mappings
        test_content = " ".join([
            test_case.get('title', ''),
            test_case.get('description', ''),
            test_case.get('steps', ''),
            test_case.get('query', '')
        ]).lower()
        
        for req in requirements:
            req_keywords = set(req['content'].lower().split())
            test_keywords = set(test_content.split())
            
            # Simple keyword overlap check
            overlap = req_keywords.intersection(test_keywords)
            if len(overlap) >= 3:  # Minimum overlap threshold
                if req['id'] not in mapped_reqs:
                    mapped_reqs.append(req['id'])
        
        return mapped_reqs
    
    def calculate_coverage_stats(self, matrix_data):
        """
        Calculate coverage statistics from matrix data.
        """
        if not matrix_data:
            return {}
        
        requirements = matrix_data['requirements']
        test_cases = matrix_data['test_cases']
        
        total_reqs = len(requirements)
        covered_reqs = sum(1 for req in requirements if req['covered'])
        uncovered_reqs = total_reqs - covered_reqs
        
        # Coverage by type
        coverage_by_type = {}
        for req in requirements:
            req_type = req['type']
            if req_type not in coverage_by_type:
                coverage_by_type[req_type] = {'total': 0, 'covered': 0}
            
            coverage_by_type[req_type]['total'] += 1
            if req['covered']:
                coverage_by_type[req_type]['covered'] += 1
        
        # Coverage by priority
        coverage_by_priority = {}
        for req in requirements:
            priority = req['priority']
            if priority not in coverage_by_priority:
                coverage_by_priority[priority] = {'total': 0, 'covered': 0}
            
            coverage_by_priority[priority]['total'] += 1
            if req['covered']:
                coverage_by_priority[priority]['covered'] += 1
        
        # Test case distribution
        tc_by_type = {}
        for tc in test_cases:
            tc_type = tc['type']
            tc_by_type[tc_type] = tc_by_type.get(tc_type, 0) + 1
        
        return {
            'overall_coverage': {
                'total_requirements': total_reqs,
                'covered_requirements': covered_reqs,
                'uncovered_requirements': uncovered_reqs,
                'coverage_percentage': round((covered_reqs / total_reqs * 100) if total_reqs > 0 else 0, 2)
            },
            'coverage_by_type': {
                req_type: {
                    'total': data['total'],
                    'covered': data['covered'],
                    'percentage': round((data['covered'] / data['total'] * 100) if data['total'] > 0 else 0, 2)
                }
                for req_type, data in coverage_by_type.items()
            },
            'coverage_by_priority': {
                priority: {
                    'total': data['total'],
                    'covered': data['covered'],
                    'percentage': round((data['covered'] / data['total'] * 100) if data['total'] > 0 else 0, 2)
                }
                for priority, data in coverage_by_priority.items()
            },
            'test_case_distribution': tc_by_type,
            'uncovered_requirements': [
                {'id': req['id'], 'content': req['content'][:100] + "..."}
                for req in requirements if not req['covered']
            ][:10]  # Top 10 uncovered
        }
    
    def export_to_excel(self, matrix_data, filename=None):
        """
        Export traceability matrix to Excel file.
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'traceability_matrix_{timestamp}.xlsx'
        
        filepath = os.path.join('data', 'exports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        workbook = Workbook()
        
        # Create worksheets
        self._create_matrix_sheet(workbook, matrix_data)
        self._create_requirements_sheet(workbook, matrix_data)
        self._create_test_cases_sheet(workbook, matrix_data)
        self._create_coverage_sheet(workbook, matrix_data)
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        workbook.save(filepath)
        return filepath
    
    def _create_matrix_sheet(self, workbook, matrix_data):
        """
        Create main traceability matrix sheet.
        """
        ws = workbook.active
        ws.title = "Traceability Matrix"
        
        requirements = matrix_data['requirements']
        test_cases = matrix_data['test_cases']
        mappings = matrix_data['mappings']
        
        # Headers
        ws['A1'] = "Requirement ID"
        ws['B1'] = "Requirement Type"
        ws['C1'] = "Priority"
        ws['D1'] = "Content"
        
        # Test case headers (starting from column E)
        start_col = 5  # Column E
        for i, tc in enumerate(test_cases):
            col_letter = get_column_letter(start_col + i)
            ws[f'{col_letter}1'] = tc['id']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, start_col + len(test_cases)):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fill requirement data
        for row, req in enumerate(requirements, start=2):
            ws[f'A{row}'] = req['id']
            ws[f'B{row}'] = req['type']
            ws[f'C{row}'] = req['priority']
            ws[f'D{row}'] = req['content']
            
            # Mark test case mappings
            req_id = req['id']
            mapped_tcs = mappings.get(req_id, [])
            
            for col, tc in enumerate(test_cases, start=start_col):
                cell = ws.cell(row=row, column=col)
                if tc['id'] in mapped_tcs:
                    cell.value = "✓"
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_requirements_sheet(self, workbook, matrix_data):
        """
        Create detailed requirements sheet.
        """
        ws = workbook.create_sheet("Requirements Details")
        
        headers = ["ID", "Type", "Priority", "Category", "Content", "Covered", "Test Case Count"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Style headers
        self._style_headers(ws, len(headers))
        
        # Fill data
        for row, req in enumerate(matrix_data['requirements'], start=2):
            ws.cell(row=row, column=1, value=req['id'])
            ws.cell(row=row, column=2, value=req['type'])
            ws.cell(row=row, column=3, value=req['priority'])
            ws.cell(row=row, column=4, value=req['category'])
            ws.cell(row=row, column=5, value=req['content'])
            ws.cell(row=row, column=6, value="Yes" if req['covered'] else "No")
            ws.cell(row=row, column=7, value=req['test_case_count'])
            
            # Color-code coverage
            coverage_cell = ws.cell(row=row, column=6)
            if req['covered']:
                coverage_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                coverage_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        self._auto_adjust_columns(ws)
    
    def _create_test_cases_sheet(self, workbook, matrix_data):
        """
        Create detailed test cases sheet.
        """
        ws = workbook.create_sheet("Test Cases Details")
        
        headers = ["ID", "Title", "Type", "Priority", "Status", "Mapped Requirements"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self._style_headers(ws, len(headers))
        
        # Fill data
        for row, tc in enumerate(matrix_data['test_cases'], start=2):
            ws.cell(row=row, column=1, value=tc['id'])
            ws.cell(row=row, column=2, value=tc['title'])
            ws.cell(row=row, column=3, value=tc['type'])
            ws.cell(row=row, column=4, value=tc['priority'])
            ws.cell(row=row, column=5, value=tc['status'])
            ws.cell(row=row, column=6, value=", ".join(tc['requirement_mappings']))
        
        self._auto_adjust_columns(ws)
    
    def _create_coverage_sheet(self, workbook, matrix_data):
        """
        Create coverage statistics sheet.
        """
        ws = workbook.create_sheet("Coverage Statistics")
        
        coverage_stats = self.calculate_coverage_stats(matrix_data)
        
        row = 1
        
        # Overall coverage
        ws[f'A{row}'] = "Overall Coverage"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        overall = coverage_stats['overall_coverage']
        ws[f'A{row}'] = "Total Requirements:"
        ws[f'B{row}'] = overall['total_requirements']
        row += 1
        
        ws[f'A{row}'] = "Covered Requirements:"
        ws[f'B{row}'] = overall['covered_requirements']
        row += 1
        
        ws[f'A{row}'] = "Uncovered Requirements:"
        ws[f'B{row}'] = overall['uncovered_requirements']
        row += 1
        
        ws[f'A{row}'] = "Coverage Percentage:"
        ws[f'B{row}'] = f"{overall['coverage_percentage']}%"
        row += 3
        
        # Coverage by type
        ws[f'A{row}'] = "Coverage by Type"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Type"
        ws[f'B{row}'] = "Total"
        ws[f'C{row}'] = "Covered"
        ws[f'D{row}'] = "Percentage"
        row += 1
        
        for req_type, data in coverage_stats['coverage_by_type'].items():
            ws[f'A{row}'] = req_type
            ws[f'B{row}'] = data['total']
            ws[f'C{row}'] = data['covered']
            ws[f'D{row}'] = f"{data['percentage']}%"
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _style_headers(self, ws, num_cols):
        """
        Apply header styling.
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _auto_adjust_columns(self, ws):
        """
        Auto-adjust column widths.
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
